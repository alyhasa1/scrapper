"""Utilities to verify eBay listing variant stock statuses.

This script reads variants from the provided Excel sheet, fetches each listing
URL, and attempts to determine whether the variant is in stock by inspecting the
resulting HTML with BeautifulSoup.

Usage:
    python ebay_stock_scraper.py --excel "ECOMWITH YASIR STOCK PRODUCTS SHEET (1).xlsx"

Optional arguments allow configuring output paths, request delay, and cookie
headers for navigating eBay bot protection.
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib import request as urllib_request
from urllib.parse import parse_qs, urlparse

import pandas as pd
from bs4 import BeautifulSoup

try:  # pragma: no cover - gracefully handle missing dependency
    import requests
except ImportError:  # pragma: no cover - fallback path is exercised in runtime
    requests = None  # type: ignore

try:  # pragma: no cover - optional dependency for headful runs
    from playwright.sync_api import (  # type: ignore
        TimeoutError as PlaywrightTimeoutError,
        sync_playwright,
    )
except ImportError:  # pragma: no cover
    sync_playwright = None  # type: ignore
    PlaywrightTimeoutError = Exception  # type: ignore


if requests is not None:  # pragma: no cover - type alias only
    HTTPErrorBase = requests.HTTPError
else:
    class HTTPErrorBase(RuntimeError):
        """Lightweight HTTPError replacement when requests isn't available."""


@dataclass
class SimpleResponse:
    """Minimal response object for urllib fallback."""

    status_code: int
    text: str


class UrllibSession:
    """Session-like wrapper to mirror requests.Session for urllib usage."""

    def __init__(self, headers: Dict[str, str], cookie_header: Optional[str]) -> None:
        self.headers = headers
        self.cookie_header = cookie_header

    def get(self, url: str, timeout: float = 30.0) -> SimpleResponse:
        req = urllib_request.Request(url, headers=self.headers)
        if self.cookie_header:
            req.add_header("Cookie", self.cookie_header)
        with urllib_request.urlopen(req, timeout=timeout) as resp:
            status = getattr(resp, "status", 200)
            body = resp.read().decode("utf-8", errors="ignore")
        return SimpleResponse(status_code=status, text=body)

LOGGER = logging.getLogger("ebay_stock_scraper")


DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-GB,en-US;q=0.9,en;q=0.8",
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp"
        ",image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7"
    ),
    "Connection": "keep-alive",
    "Cache-Control": "no-cache",
}

CHALLENGE_MARKERS: Iterable[str] = (
    "Pardon our interruption",
    "captcha",
    "ebay-static-captcha",
    "challenge",
)

CHALLENGE_MARKERS_LOWER: Tuple[str, ...] = tuple(marker.lower() for marker in CHALLENGE_MARKERS)

AVAILABILITY_JSON_KEYS: Iterable[str] = (
    "availability",
    "availabilityType",
)

AVAILABILITY_STRINGS = {
    "http://schema.org/InStock": "IN_STOCK",
    "http://schema.org/OutOfStock": "OUT_OF_STOCK",
    "InStock": "IN_STOCK",
    "OutOfStock": "OUT_OF_STOCK",
    "AVAILABLE": "IN_STOCK",
    "UNAVAILABLE": "OUT_OF_STOCK",
}


@dataclass
class VariantRecord:
    """Represents a single variant row from the spreadsheet."""

    row_index: int
    excel_row: int
    source_url: str
    sr_no: Optional[str]
    stock_name: Optional[str]
    variation: Optional[str]
    size: Optional[str]
    dimensions: Optional[str]
    sheet_stock_status: Optional[str]
    item_number: Optional[str]
    listing_url: str

    @property
    def base_item_id(self) -> Optional[str]:
        if not self.listing_url:
            return None
        parsed = urlparse(self.listing_url)
        path_parts = parsed.path.rstrip("/").split("/")
        try:
            return next(part for part in reversed(path_parts) if part.isdigit())
        except StopIteration:
            return None

    @property
    def variation_id(self) -> Optional[str]:
        parsed = urlparse(self.source_url)
        query = parse_qs(parsed.query)
        var_values = query.get("var")
        if var_values:
            return var_values[0]
        return None


def parse_cookie_header(cookie_header: Optional[str]) -> Dict[str, str]:
    if not cookie_header:
        return {}
    pairs: Dict[str, str] = {}
    for chunk in cookie_header.split(";"):
        if "=" not in chunk:
            continue
        name, value = chunk.split("=", 1)
        pairs[name.strip()] = value.strip()
    return pairs


def configure_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s - %(levelname)s - %(message)s",
        datefmt="%H:%M:%S",
    )


def load_variants_from_excel(
    path: Path,
    limit: Optional[int] = None,
    start_row: int = 1,
) -> List[VariantRecord]:
    if not path.exists():
        raise FileNotFoundError(path)

    df = pd.read_excel(path)
    if df.empty:
        return []

    if start_row is None or start_row < 1:
        start_row = 1

    header_row = df.iloc[0]
    resolved_headers = []
    for idx, col in enumerate(df.columns):
        candidate = header_row.iloc[idx]
        if isinstance(candidate, str) and candidate.strip():
            resolved_headers.append(candidate.strip())
        else:
            resolved_headers.append(col)
    df = df.iloc[1:].reset_index(drop=True)
    df.columns = resolved_headers

    def clean_cell_value(value: Any) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, str):
            cleaned = value.strip()
            return cleaned or None
        try:
            if pd.isna(value):  # type: ignore[arg-type]
                return None
        except TypeError:
            pass
        text = str(value).strip()
        return text or None

    column_map = {
        "Sr No": "sr_no",
        "account Sku Number": "sr_no",
        "Product Name": "stock_name",
        "STOCK NAME": "stock_name",
        "Color / Design": "variation",
        "Colour": "variation",
        "Color": "variation",
        "Pack/TYPE": "pack_type",
        "Size / Pack/TYPE": "pack_type",
        "DIMENSIONS": "dimensions",
        "SIZE": "dimensions",
        "STATUS": "sheet_stock_status",
        "INSTOCK/OUTOFSTOCK": "sheet_stock_status",
        "ITEM NUMBER": "item_number",
        "LISTING LINK": "listing_url",
    }

    df = df.rename(columns=column_map)

    if "listing_url" not in df.columns:
        raise ValueError("Expected 'LISTING LINK' column to be present in the sheet.")

    df = df[df["listing_url"].notna()].copy()
    df["listing_url"] = df["listing_url"].astype(str).str.strip()

    records: List[VariantRecord] = []
    history_records: List[VariantRecord] = []
    last_variation_by_listing: Dict[str, Optional[str]] = {}
    for idx, row in enumerate(df.itertuples(index=False), start=1):
        raw_url = getattr(row, "listing_url")
        parsed_url = urlparse(raw_url)
        base_url = parsed_url._replace(query="", fragment="").geturl()
        color = clean_cell_value(getattr(row, "variation", None))
        if color:
            last_variation_by_listing[base_url] = color
        else:
            color = last_variation_by_listing.get(base_url)
        pack_type = clean_cell_value(getattr(row, "pack_type", None))
        dimensions = clean_cell_value(getattr(row, "dimensions", None))
        status_value = clean_cell_value(getattr(row, "sheet_stock_status", None))
        product_name = clean_cell_value(getattr(row, "stock_name", None))
        
        # If no color found directly, try to extract from product name
        if not color and product_name:
            parts = product_name.split()
            for i, part in enumerate(parts):
                if part.lower() in ["cream", "black", "grey", "gray", "brown", "beige", "blue", "green", "white", "red", "silver", "dark"]:
                    remaining = " ".join(parts[i:])
                    if " - " in remaining:
                        color = remaining.split(" - ")[0].strip() + " - Greekey"
                    else:
                        color = remaining.strip()
                    LOGGER.debug("Extracted color from product name: %s", color)
                    break
        
        # Handle data inheritance: if no color found, check if we can inherit from previous rows
        # with the same base URL and a similar color pattern
        if not color:
            current_size = dimensions or pack_type
            if current_size:
                # Look back through previous rows with the same base listing URL
                for prev_idx in range(idx - 1, 0, -1):
                    prev_record = history_records[prev_idx - 1]
                    if prev_record.listing_url == base_url:
                        # Check if previous record has a valid color that matches the pattern
                        if prev_record.variation and " - Greekey" in prev_record.variation:
                            # Check if sizes are related (same product, different sizes)
                            current_size_clean = current_size.lower().replace(" cm", "").replace("cm", "").strip()
                            prev_size_clean = prev_record.size.lower().replace(" cm", "").replace("cm", "").strip() if prev_record.size else ""
                            
                            # If previous record has a size (likely first row in color group),
                            # inherit the color
                            if prev_size_clean:
                                color = prev_record.variation
                                LOGGER.debug("Inherited color '%s' from previous row %d for size %s",
                                           color, prev_record.excel_row, current_size)
                                break
                        elif prev_record.stock_name and not color:
                            # Try to extract color from previous record's product name
                            prev_parts = prev_record.stock_name.split()
                            for i, part in enumerate(prev_parts):
                                if part.lower() in ["cream", "black", "grey", "gray", "brown", "beige", "blue", "green", "white", "red", "silver", "dark"]:
                                    remaining = " ".join(prev_parts[i:])
                                    if " - " in remaining:
                                        inherited_color = remaining.split(" - ")[0].strip() + " - Greekey"
                                    else:
                                        inherited_color = remaining.strip()
                                    color = inherited_color
                                    LOGGER.debug("Extracted inherited color '%s' from previous row %d product name", 
                                               color, prev_record.excel_row)
                                    break
                            if color:
                                break
                    # If we hit a different product (different URL), stop looking back
                    elif prev_record.listing_url != base_url:
                        break

        size_value = dimensions or pack_type
        if size_value and isinstance(size_value, str) and size_value.strip().lower() in {"", "n/a", "na", "select", "rug"}:
            size_value = dimensions
        if not size_value:
            size_value = pack_type
        record = VariantRecord(
            row_index=idx,
            excel_row=idx + 2,  # account for excel header row + header data row
            source_url=raw_url,
            sr_no=getattr(row, "sr_no", None),
            stock_name=product_name,
            variation=color,
            size=size_value,
            dimensions=dimensions,
            sheet_stock_status=status_value,
            item_number=str(getattr(row, "item_number", "")) if getattr(row, "item_number", None) else None,
            listing_url=base_url,
        )
        history_records.append(record)
        if idx < start_row:
            continue
        records.append(record)
        if limit is not None and len(records) >= limit:
            break
    return records


def create_session(headers: Dict[str, str], cookie_header: Optional[str]) -> Any:
    if requests is None:
        return UrllibSession(headers=headers, cookie_header=cookie_header)

    session = requests.Session()
    session.headers.update(headers)
    cookies = parse_cookie_header(cookie_header)
    if cookies:
        session.cookies.update(cookies)
    return session


class ChallengeDetected(RuntimeError):
    """Raised when bot protection content is received."""


def fetch_listing_html(session: Any, url: str, timeout: float = 30.0) -> str:
    LOGGER.debug("Fetching %s", url)
    response = session.get(url, timeout=timeout)
    if response.status_code != 200:
        raise HTTPErrorBase(f"Unexpected status {response.status_code} for {url}")
    text = response.text
    lowered = text.lower()
    if any(marker in lowered for marker in CHALLENGE_MARKERS_LOWER):
        raise ChallengeDetected(f"Bot challenge detected for {url}")
    return text


def parse_availability_from_json(data: Dict[str, object]) -> Optional[str]:
    for key in AVAILABILITY_JSON_KEYS:
        value = data.get(key)
        if isinstance(value, str):
            standardised = AVAILABILITY_STRINGS.get(value)
            if standardised:
                return standardised
    if "offers" in data:
        offers = data["offers"]
        if isinstance(offers, dict):
            availability = offers.get("availability")
            if isinstance(availability, str):
                standardised = AVAILABILITY_STRINGS.get(availability)
                if standardised:
                    return standardised
    return None


def detect_availability(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")

    # Prefer structured data blocks.
    for script in soup.find_all("script", attrs={"type": "application/ld+json"}):
        try:
            payload = json.loads(script.string)
        except (TypeError, json.JSONDecodeError):
            continue
        if isinstance(payload, dict):
            availability = parse_availability_from_json(payload)
            if availability:
                return availability
        elif isinstance(payload, list):
            for entry in payload:
                if isinstance(entry, dict):
                    availability = parse_availability_from_json(entry)
                    if availability:
                        return availability

    # Fallback: look for explicit text markers.
    normalized = soup.get_text(" ", strip=True).lower()
    if "out of stock" in normalized:
        return "OUT_OF_STOCK"
    if "in stock" in normalized or "last one" in normalized:
        return "IN_STOCK"

    # Fallback: search within data attributes.
    if "\"availability\":" in html:
        for token, status in AVAILABILITY_STRINGS.items():
            if token in html:
                return status

    return "UNKNOWN"


def process_variants_requests(
    records: List[VariantRecord],
    session: Any,
    delay: float,
    max_retries: int,
) -> pd.DataFrame:
    results = []
    for index, record in enumerate(records, start=1):
        LOGGER.info(
            "[%s/%s] Checking item %s (variant %s)",
            index,
            len(records),
            record.item_number or record.base_item_id,
            record.variation or record.size or record.variation_id,
        )

        attempts = 0
        availability = "BLOCKED"
        error_message = None
        while attempts <= max_retries:
            attempts += 1
            try:
                html = fetch_listing_html(session, record.listing_url)
                availability = detect_availability(html)
                error_message = None
                break
            except ChallengeDetected as exc:
                error_message = str(exc)
                LOGGER.warning("Challenge detected for %s", record.listing_url)
                availability = "BLOCKED"
                time.sleep(delay * 2)
            except Exception as exc:  # pylint: disable=broad-except
                error_message = str(exc)
                LOGGER.exception("Failed to process %s", record.listing_url)
                availability = "ERROR"
                time.sleep(delay)
        else:
            LOGGER.error(
                "Max retries exceeded for %s", record.listing_url
            )

        results.append(
            {
                "row_index": record.row_index,
                "excel_row": record.excel_row,
                "sheet_sr_no": record.sr_no,
                "stock_name": record.stock_name,
                "variation": record.variation,
                "size": record.size,
                "dimensions": record.dimensions,
                "sheet_stock_status": record.sheet_stock_status,
                "detected_status": availability,
                "item_number": record.item_number or record.base_item_id,
                "variation_id": record.variation_id,
                "listing_url": record.listing_url,
                "source_url": record.source_url,
                "error": error_message,
            }
        )

        if index < len(records):
            time.sleep(delay)

    return pd.DataFrame(results)


def normalize_label(text: str) -> str:
    return " ".join(text.lower().split())


def extract_base_colour(colour_text: str) -> str:
    """Extract base colour name by removing seller-specific suffixes."""
    # Remove common suffixes like "- Greekey", "- Gel Back 59", "- MR (41)", etc.
    text = colour_text.strip()
    # Split on first dash and take the part before it
    if ' - ' in text:
        base = text.split(' - ')[0].strip()
        return base.lower()
    return text.lower()


def extract_base_size(size_text: str) -> str:
    """Extract base size by removing imperial measurements and extra text."""
    # Remove imperial measurements in parentheses like "(1 ft 4 in x 2 ft)"
    text = size_text.strip()
    # Remove everything in parentheses
    if '(' in text:
        text = text.split('(')[0].strip()
    # Remove "Most popular", "selected", etc.
    text = text.replace('Most popular', '').replace('selected', '').strip()
    return text.lower()


def option_matches(option_text: str, target: str) -> bool:
    """Match option text against target with flexible matching for colours and sizes."""
    option_normalized = normalize_label(option_text)
    target_normalized = normalize_label(target)
    
    # First try exact substring match
    if target_normalized in option_normalized:
        LOGGER.debug("Match via substring: target '%s' found in option '%s'", target_normalized, option_normalized)
        return True
    
    # For colour matching: extract base colour names and compare
    # Check if this looks like a colour option (contains " - ")
    if ' - ' in option_text or ' - ' in target:
        option_base = extract_base_colour(option_text)
        target_base = extract_base_colour(target)
        if option_base == target_base or target_base in option_base or option_base in target_base:
            LOGGER.debug("Match via colour base: option '%s' -> '%s', target '%s' -> '%s'", 
                        option_text, option_base, target, target_base)
            return True
    
    # For size matching: strip imperial measurements and compare
    # Check if this looks like a size (contains "cm" or "x" followed by digits)
    if ('cm' in option_text.lower() or ('x' in option_text.lower() and any(c.isdigit() for c in option_text))):
        option_base = extract_base_size(option_text)
        target_base = extract_base_size(target)
        # Normalize spaces around 'x' for comparison (e.g., "60x110" vs "60 x 110")
        option_base_clean = option_base.replace(' ', '')
        target_base_clean = target_base.replace(' ', '')
        if option_base == target_base or target_base in option_base or option_base_clean == target_base_clean:
            LOGGER.debug("Match via size base: option '%s' -> '%s', target '%s' -> '%s'", 
                        option_text, option_base, target, target_base)
            return True
    
    LOGGER.debug("No match: option '%s', target '%s'", option_text, target)
    return False


def interpret_option_state(option_text: str, class_attr: Optional[str], aria_disabled: Optional[str]) -> Tuple[bool, str]:
    text_lower = option_text.lower()
    disabled = False
    reason = ""
    if class_attr and "listbox__option--disabled" in class_attr:
        disabled = True
        reason = "Option marked out of stock (disabled class)"
    if aria_disabled == "true":
        disabled = True
        reason = "Option marked out of stock (aria-disabled=true)"
    if "out of stock" in text_lower:
        disabled = True
        reason = "Label indicates out of stock"
    return disabled, reason


def collect_variant_groups(page) -> List[Dict[str, Any]]:
    groups: List[Dict[str, Any]] = []

    # Primary SKU widgets
    containers = page.locator("div.vim.x-sku")
    container_count = containers.count()
    LOGGER.debug("Found %d variant containers (div.vim.x-sku)", container_count)
    for idx in range(containers.count()):
        container = containers.nth(idx)
        try:
            label = container.locator(".btn__label").inner_text(timeout=1000).strip()
        except PlaywrightTimeoutError:
            label = None
        if not label:
            continue
        LOGGER.debug("Found group '%s' via div.vim.x-sku", label.rstrip(":"))
        groups.append({
            "label": label.rstrip(":"),
            "locator": container,
            "type": "vim",
        })

    # Alternative button based selectors (data-testid)
    alt_buttons = page.locator("[data-testid='x-msku__group']")
    for idx in range(alt_buttons.count()):
        group = alt_buttons.nth(idx)
        try:
            label = group.locator("[data-testid='x-msku__group-title']").inner_text(timeout=1000).strip()
        except PlaywrightTimeoutError:
            label = None
        if not label:
            continue
        groups.append({
            "label": label.rstrip(":"),
            "locator": group,
            "type": "msku",
        })

    # Fallback to select dropdowns (older layout)
    legacy_selects = page.locator("select[name^='variation'], select#msku-sel-1")
    for idx in range(legacy_selects.count()):
        select_el = legacy_selects.nth(idx)
        label = select_el.evaluate("el => el.closest('div')?.querySelector('label, span')?.innerText || el.getAttribute('name')")
        if not label:
            continue
        groups.append({
            "label": str(label).strip().rstrip(":"),
            "locator": select_el,
            "type": "select",
        })

    LOGGER.debug("Total variant groups collected: %d", len(groups))
    for idx, grp in enumerate(groups):
        LOGGER.debug("  Group %d: '%s' (type: %s)", idx, grp["label"], grp["type"])
    
    return groups


GROUP_KEYWORDS = {
    "size": ["size", "length", "dimensions"],
    "variation": ["colour", "color", "design", "style", "pattern"],
}


def find_group_for_dimension(groups: List[Dict[str, Any]], dimension: str, used_indices: set) -> Optional[int]:
    keywords = GROUP_KEYWORDS.get(dimension, [])
    for idx, group in enumerate(groups):
        if idx in used_indices:
            continue
        label_norm = normalize_label(group["label"])
        if any(keyword in label_norm for keyword in keywords):
            return idx
    # fallback to first unused
    for idx, group in enumerate(groups):
        if idx not in used_indices:
            return idx
    return None


def select_option_from_group(page, group: Dict[str, Any], target_value: str) -> Tuple[bool, Optional[str]]:
    locator = group["locator"]
    group_type = group.get("type", "vim")
    group_label = group.get("label", "unknown")
    
    LOGGER.debug("Selecting '%s' from group '%s' (type: %s)", target_value, group_label, group_type)

    if group_type == "select":
        try:
            locator.scroll_into_view_if_needed(timeout=2000)
        except PlaywrightTimeoutError:
            pass
        select = locator
        options = select.locator("option")
        count = options.count()
        for idx in range(count):
            option = options.nth(idx)
            text = option.inner_text().strip()
            if option_matches(text, target_value):
                disabled_attr = option.get_attribute("disabled")
                if disabled_attr is not None:
                    return False, "Option disabled"
                option_value = option.get_attribute("value") or text
                select.select_option(option_value)
                page.wait_for_timeout(400)
                return True, None
        return False, f"Variant value '{target_value}' not found"

    # Button-based selection
    button_locators = [
        locator.locator("button.listbox-button__control"),
        locator.locator("button[aria-haspopup='listbox']"),
    ]
    button = None
    for candidate in button_locators:
        if candidate.count():
            button = candidate.first
            break
    if button is None:
        return False, "Variant selector button missing"

    try:
        button.scroll_into_view_if_needed(timeout=2000)
    except PlaywrightTimeoutError:
        pass
    button.click()
    page.wait_for_timeout(300)

    options_container = locator
    options_in_button = locator.locator("div.listbox__options")
    if options_in_button.count() > 0:
        try:
            options_in_button.first.wait_for(state="visible", timeout=2000)
            options_container = options_in_button.first
        except PlaywrightTimeoutError:
            pass

    if options_container == locator:
        aria_controls = button.get_attribute("aria-controls")
        if aria_controls:
            targeted = page.locator(f"#{aria_controls}")
            try:
                targeted.wait_for(state="visible", timeout=2000)
                if targeted.count():
                    options_container = targeted
            except PlaywrightTimeoutError:
                pass

    options = options_container.locator("div.listbox__option")
    if options.count() == 0:
        options = options_container.locator("li[role='option']")
    
    chosen_option = None
    chosen_text = ""
    
    for idx in range(options.count()):
        option = options.nth(idx)
        data_sku_value = option.get_attribute("data-sku-value-name")
        if data_sku_value and option_matches(data_sku_value, target_value):
            chosen_option = option
            chosen_text = option.inner_text().strip()
            LOGGER.debug("Matched option via data-sku-value-name: %s", data_sku_value)
            break
    
    if chosen_option is None:
        for idx in range(options.count()):
            option = options.nth(idx)
            text = option.inner_text().strip()
            if option_matches(text, target_value):
                chosen_option = option
                chosen_text = text
                LOGGER.debug("Matched option via text: %s", text)
                break
    if chosen_option is None:
        button.press("Escape")
        return False, f"Variant value '{target_value}' not found"

    class_attr = chosen_option.get_attribute("class")
    aria_disabled = chosen_option.get_attribute("aria-disabled")
    LOGGER.debug("Option state check - text: %s, class: %s, aria-disabled: %s", chosen_text, class_attr, aria_disabled)
    disabled, reason = interpret_option_state(chosen_text, class_attr, aria_disabled)
    if disabled:
        LOGGER.info("Option is disabled: %s", reason)
        button.press("Escape")
        return False, reason or "Option disabled"

    # Scroll option into view within dropdown and wait for it to be visible
    try:
        chosen_option.scroll_into_view_if_needed(timeout=3000)
        chosen_option.wait_for(state="visible", timeout=3000)
    except PlaywrightTimeoutError:
        LOGGER.warning("Option not visible after scroll, attempting click anyway")
    
    try:
        chosen_option.click(timeout=10000)
    except PlaywrightTimeoutError:
        button.press("Escape")
        return False, f"Timeout clicking option '{target_value}' - element not interactable"
    
    page.wait_for_timeout(800)
    return True, None


def evaluate_availability(page) -> Tuple[str, Optional[str]]:
    script = """
    () => {
      const selectors = [
        '[data-testid="x-msku__availability-message"]',
        '#x-msku__availability-message',
        '#qtySubTxt',
        '[data-testid="availability-text"]',
        '[data-testid="availability-messaging"]'
      ];
      for (const sel of selectors) {
        const el = document.querySelector(sel);
        if (el && el.textContent) {
          const text = el.textContent.trim();
          if (text) return text;
        }
      }
      const button = document.querySelector('button[data-testid="art-atc-button"]') || document.querySelector('#atcRedesignId_btn');
      if (button) {
        const disabled = button.getAttribute('aria-disabled') === 'true' || button.hasAttribute('disabled');
        return disabled ? 'Add to basket disabled' : 'Add to basket enabled';
      }
      return null;
    }
    """
    try:
        availability_text = page.evaluate(script)
    except Exception as exc:
        LOGGER.warning("Failed to evaluate availability script, retrying after delay: %s", exc)
        page.wait_for_timeout(1000)
        try:
            availability_text = page.evaluate(script)
        except Exception as retry_exc:
            LOGGER.error("Failed to evaluate availability after retry: %s", retry_exc)
            return "IN_STOCK", None
    
    if availability_text:
        lowered = availability_text.lower()
        if "out of stock" in lowered or "unavailable" in lowered:
            return "OUT_OF_STOCK", availability_text
        if "add to basket disabled" in lowered:
            return "OUT_OF_STOCK", availability_text
    return "IN_STOCK", availability_text


def detect_challenge(page) -> bool:
    try:
        main_text = page.inner_text("body", timeout=2000).lower()
    except PlaywrightTimeoutError:
        main_text = ""
    if any(marker in main_text for marker in CHALLENGE_MARKERS_LOWER):
        return True
    if page.url and "splashui/challenge" in page.url:
        return True
    return False


def process_variants_chromium(
    records: List[VariantRecord],
    delay: float,
    max_retries: int,
    cookie_header: Optional[str],
    user_data_dir: Optional[Path],
) -> pd.DataFrame:
    if sync_playwright is None:
        raise RuntimeError("Playwright is not installed. Run 'pip install playwright' and 'playwright install chromium'.")

    results: List[Dict[str, Any]] = []
    with sync_playwright() as p:
        browser = None
        context = None
        if user_data_dir:
            context = p.chromium.launch_persistent_context(
                str(user_data_dir),
                headless=False,
                viewport={"width": 1366, "height": 900},
                user_agent=DEFAULT_HEADERS["User-Agent"],
                locale="en-GB",
                accept_downloads=False,
                ignore_https_errors=True,
            )
        else:
            browser = p.chromium.launch(headless=False)
            context = browser.new_context(
                viewport={"width": 1366, "height": 900},
                user_agent=DEFAULT_HEADERS["User-Agent"],
                locale="en-GB",
                accept_downloads=False,
                ignore_https_errors=True,
            )

        assert context is not None  # narrow type checker
        context.set_default_timeout(45000)
        context.set_default_navigation_timeout(45000)

        if cookie_header:
            cookies = []
            for chunk in cookie_header.split(";"):
                chunk = chunk.strip()
                if not chunk or "=" not in chunk:
                    continue
                name, value = chunk.split("=", 1)
                cookies.append({
                    "name": name.strip(),
                    "value": value.strip(),
                    "domain": ".ebay.co.uk",
                    "path": "/",
                })
            if cookies:
                try:
                    context.add_cookies(cookies)
                except Exception as exc:  # pragma: no cover
                    LOGGER.warning("Failed to set cookies: %s", exc)

        page = context.new_page()
        for index, record in enumerate(records, start=1):
            LOGGER.info("[Chromium %s/%s] Checking item %s", index, len(records), record.item_number or record.base_item_id)
            attempts = 0
            availability = "BLOCKED"
            error_message: Optional[str] = None
            while attempts <= max_retries:
                attempts += 1
                try:
                    page.goto(
                        record.listing_url,
                        wait_until="domcontentloaded",
                        timeout=45000,
                    )
                    try:
                        page.wait_for_load_state("networkidle", timeout=5000)
                    except PlaywrightTimeoutError:
                        pass
                    page.wait_for_timeout(800)
                    if detect_challenge(page):
                        raise ChallengeDetected("Bot challenge page encountered")

                    groups = collect_variant_groups(page)
                    used_indices: set = set()
                    dimension_results: List[str] = []

                    # Build dimension order dynamically based on actual group order on page
                    # First, find which group corresponds to each dimension
                    dimensions_with_group: List[Tuple[int, str, str]] = []  # (group_idx, dim_type, dim_value)
                    
                    if record.variation:
                        group_idx = find_group_for_dimension(groups, "variation", set())
                        if group_idx is not None:
                            dimensions_with_group.append((group_idx, "variation", str(record.variation)))
                            LOGGER.debug("Found colour group at index %d: %s", group_idx, record.variation)
                        else:
                            LOGGER.debug("No colour group found for row %s (item %s)", record.row_index, record.item_number or record.base_item_id)
                    else:
                        LOGGER.debug("No colour provided for row %s (item %s)", record.row_index, record.item_number or record.base_item_id)
                    
                    if record.size:
                        group_idx = find_group_for_dimension(groups, "size", set())
                        if group_idx is not None:
                            dimensions_with_group.append((group_idx, "size", str(record.size)))
                            LOGGER.debug("Found size group at index %d: %s", group_idx, record.size)
                        else:
                            LOGGER.debug("No size group found for row %s (item %s)", record.row_index, record.item_number or record.base_item_id)
                    else:
                        LOGGER.debug("No size provided for row %s (item %s)", record.row_index, record.item_number or record.base_item_id)
                    
                    # Sort by group index to process in page order (color-first or size-first)
                    dimensions_with_group.sort(key=lambda x: x[0])
                    LOGGER.debug("Processing dimensions in page order: %s", [(dt, dv) for _, dt, dv in dimensions_with_group])

                    selection_failed = False
                    for group_idx, dim_type, dim_value in dimensions_with_group:
                        if group_idx in used_indices:
                            LOGGER.warning("Group %d already used, skipping", group_idx)
                            continue
                        used_indices.add(group_idx)
                        group = groups[group_idx]
                        ok, failure_reason = select_option_from_group(page, group, dim_value)
                        if not ok:
                            reason_lower = (failure_reason or "").lower()
                            if any(keyword in reason_lower for keyword in ["out of stock", "disabled", "aria-disabled"]):
                                availability = "OUT_OF_STOCK"
                            else:
                                availability = "ERROR"
                            error_message = failure_reason
                            selection_failed = True
                            break
                        dimension_results.append(dim_value)

                    if selection_failed:
                        break

                    try:
                        page.wait_for_load_state("networkidle", timeout=5000)
                    except PlaywrightTimeoutError:
                        LOGGER.debug("Network not idle after variant selection, continuing anyway")
                    
                    page.wait_for_timeout(1000)
                    
                    try:
                        page.wait_for_load_state("domcontentloaded", timeout=3000)
                    except PlaywrightTimeoutError:
                        pass

                    availability, availability_detail = evaluate_availability(page)
                    error_message = availability_detail if availability == "OUT_OF_STOCK" and availability_detail else None
                    break
                except ChallengeDetected as exc:
                    error_message = str(exc)
                    availability = "BLOCKED"
                    LOGGER.warning("Challenge detected during Chromium check for %s", record.listing_url)
                    page.wait_for_timeout(delay * 1000)
                except PlaywrightTimeoutError as exc:
                    error_message = f"Timeout: {exc}"
                    availability = "ERROR"
                    LOGGER.exception("Timeout for %s", record.listing_url)
                    page.wait_for_timeout(delay * 1000)
                except Exception as exc:  # pylint: disable=broad-except
                    error_message = str(exc)
                    availability = "ERROR"
                    LOGGER.exception("Chromium processing failed for %s", record.listing_url)
                    page.wait_for_timeout(delay * 1000)
            results.append(
                {
                    "row_index": record.row_index,
                    "excel_row": record.excel_row,
                    "sheet_sr_no": record.sr_no,
                    "stock_name": record.stock_name,
                    "variation": record.variation,
                    "size": record.size,
                    "dimensions": record.dimensions,
                    "sheet_stock_status": record.sheet_stock_status,
                    "detected_status": availability,
                    "item_number": record.item_number or record.base_item_id,
                    "variation_id": record.variation_id,
                    "listing_url": record.listing_url,
                    "error": error_message,
                }
            )
            if index < len(records):
                page.wait_for_timeout(delay * 1000)

        if browser is not None:
            browser.close()
        else:
            context.close()

    return pd.DataFrame(results)


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Verify eBay listing variant stock status.")
    parser.add_argument(
        "--excel",
        type=Path,
        required=True,
        help="Path to the Excel sheet containing listing data.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Limit the number of listing rows to process.",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=1,
        help="1-based row index (excluding header rows) to start processing from.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Optional path to write the results as an Excel or CSV file.",
    )
    parser.add_argument(
        "--copy-sheet",
        type=Path,
        help="Optional path to write a copy of the input sheet with updated stock statuses.",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=3.0,
        help="Delay in seconds between requests (default: 3.0).",
    )
    parser.add_argument(
        "--cookie",
        type=str,
        default=None,
        help="Optional Cookie header string to help bypass bot protection.",
    )
    parser.add_argument(
        "--retries",
        type=int,
        default=2,
        help="Number of retries when encountering errors or challenges (default: 2).",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable verbose logging output.",
    )
    parser.add_argument(
        "--engine",
        choices=["requests", "chromium"],
        default="requests",
        help="Choose backend engine: 'requests' (default) or 'chromium' for headful Playwright run.",
    )
    parser.add_argument(
        "--chromium-profile",
        type=Path,
        default=None,
        help="Optional directory for persistent Chromium user data (helps retain login).",
    )
    return parser.parse_args(argv)


def update_sheet_with_results(input_path: Path, output_path: Path, results_df: pd.DataFrame) -> None:
    from datetime import datetime
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string
    from shutil import copy2

    status_normalization = {
        "IN_STOCK": "INSTOCK",
        "OUT_OF_STOCK": "OUT OF STOCK",
        "BLOCKED": "BLOCKED",
        "ERROR": "ERROR",
        "UNKNOWN": "UNKNOWN",
    }

    status_by_row = {}
    for _, row in results_df.iterrows():
        excel_row = int(row["excel_row"])
        detected = row["detected_status"]
        if isinstance(detected, str):
            detected_clean = status_normalization.get(detected, detected.replace("_", " "))
        else:
            detected_clean = detected
        status_by_row[excel_row] = detected_clean

    # Prepare backup directory
    backup_dir = input_path.parent / "backups"
    backup_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = backup_dir / f"{input_path.stem}_{timestamp}{input_path.suffix}"
    copy2(input_path, backup_path)

    preferred_status_letter = "I"

    # Helper to resolve the STATUS column dynamically (handles column shifts)
    def resolve_status_column(ws):
        header_candidates = {"STATUS", "INSTOCK/OUTOFSTOCK"}
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, str) and cell.value.strip().upper() in header_candidates:
                    return cell.column
        # Fallback to explicit column letter if headers have been renamed
        return column_index_from_string(preferred_status_letter)

    # Update original file in place
    original_wb = load_workbook(input_path)
    original_ws = original_wb.active
    status_col_index = resolve_status_column(original_ws)
    for excel_row, status in status_by_row.items():
        original_ws.cell(row=excel_row, column=status_col_index).value = status
    original_wb.save(input_path)
    original_wb.close()

    copy2(input_path, output_path)

    # Re-open the saved copy and update only the resolved STATUS column (defaults to column I)
    copy_wb = load_workbook(output_path)
    copy_ws = copy_wb.active
    status_col_index = resolve_status_column(copy_ws)

    for excel_row, status in status_by_row.items():
        copy_ws.cell(row=excel_row, column=status_col_index).value = status

    copy_wb.save(output_path)
    copy_wb.close()


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    configure_logging(args.verbose)

    try:
        records = load_variants_from_excel(
            args.excel,
            limit=args.limit,
            start_row=args.start_row,
        )
    except Exception as exc:  # pylint: disable=broad-except
        LOGGER.error("Failed to load spreadsheet: %s", exc)
        return 1

    if not records:
        LOGGER.warning("No listing rows found in spreadsheet.")
        return 0

    if args.engine == "chromium":
        results_df = process_variants_chromium(records, args.delay, args.retries, args.cookie, args.chromium_profile)
    else:
        session = create_session(DEFAULT_HEADERS.copy(), args.cookie)
        results_df = process_variants_requests(records, session, args.delay, args.retries)

    sample = results_df.head()
    LOGGER.info("Sample results:\n%s", sample)

    if args.output:
        output_path = args.output
        if output_path.suffix.lower() in {".xlsx", ".xls"}:
            results_df.to_excel(output_path, index=False)
        else:
            results_df.to_csv(output_path, index=False)
        LOGGER.info("Results saved to %s", output_path)
    else:
        print(results_df.to_string(index=False))

    if args.copy_sheet:
        try:
            update_sheet_with_results(args.excel, args.copy_sheet, results_df)
            LOGGER.info("Updated copy of sheet saved to %s", args.copy_sheet)
        except Exception as exc:  # pylint: disable=broad-except
            LOGGER.error("Failed to write updated sheet: %s", exc)

    if args.engine != "chromium":
        blocked = results_df[results_df["detected_status"] == "BLOCKED"]
        if not blocked.empty:
            LOGGER.warning(
                "%s variants could not be checked due to bot protection. Consider switching to --engine chromium and supplying a Cookie header.",
                len(blocked),
            )

    return 0


if __name__ == "__main__":  # pragma: no cover
    sys.exit(main())
