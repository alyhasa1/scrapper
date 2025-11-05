from openpyxl import load_workbook

wb = load_workbook('sheet_rows99-103_final.xlsx')
ws = wb.active

print("Checking rows around 99-104:")
for row_num in range(99, 105):
    col_d = ws.cell(row_num, 4).value  # Colour column
    col_i = ws.cell(row_num, 9).value  # STATUS column
    print(f"Row {row_num}: Colour={col_d}, STATUS={col_i}")
