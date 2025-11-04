import itertools
import string

import pandas as pd
from openpyxl import load_workbook

path = r"d:\stock check scrapper\ECOMWITH YASIR STOCK PRODUCTS SHEET (1).xlsx"

# OpenPyXL view for raw rows --------------------------------------------------
wb = load_workbook(path, read_only=True, data_only=True)
ws = wb.active

print("Raw worksheet preview (first 12 columns, first 8 rows):")
cols = list(string.ascii_uppercase[:12]) + ["L", "M", "N", "O", "P"]
for row_idx, row in enumerate(itertools.islice(ws.iter_rows(values_only=True), 8), start=1):
    row_values = []
    for col_idx, value in enumerate(row, start=1):
        # limit preview to first 16 columns to keep output manageable
        if col_idx > 16:
            break
        letters = []
        idx_tmp = col_idx
        while idx_tmp:
            idx_tmp, rem = divmod(idx_tmp - 1, 26)
            letters.append(string.ascii_uppercase[rem])
        column_letter = "".join(reversed(letters))
        row_values.append(f"{column_letter}:{value}")
    print(f"Row {row_idx}: {row_values}")

wb.close()

# Pandas-based structure detection -------------------------------------------
df = pd.read_excel(path, header=None)

header_idx = None
for idx, row in df.iterrows():
    values = row.fillna("").astype(str).str.strip()
    if "LISTING NUMBER" in values.values and "LISTING LINK" in values.values:
        header_idx = idx
        break

if header_idx is None:
    raise RuntimeError("Unable to locate header row containing key labels")

headers = df.iloc[header_idx]

print(f"\nDetected header row index: {header_idx}")
print("Column overview (index -> letter -> header):")
for idx, value in enumerate(headers, start=1):
    letters = []
    col_idx = idx
    while col_idx:
        col_idx, rem = divmod(col_idx - 1, 26)
        letters.append(string.ascii_uppercase[rem])
    column_letter = "".join(reversed(letters))
    print(f"{idx:>02} -> {column_letter}: {value}")

print("\nFirst 8 data rows after header:")
data_rows = df.iloc[header_idx + 1: header_idx + 9]
print(data_rows)

print("\nRow dict samples (non-empty fields):")
for idx, row in data_rows.iterrows():
    row_dict = {headers[i]: row[i] for i in range(len(headers)) if pd.notna(row[i])}
    print(idx, row_dict)

