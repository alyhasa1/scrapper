from openpyxl import load_workbook
from pathlib import Path

source = Path("ECOMWITH YASIR STOCK PRODUCTS SHEET (1).xlsx")
copy_path = Path("sheet_top10_updated.xlsx")

src_wb = load_workbook(source, read_only=True)
src_ws = src_wb.active
copy_wb = load_workbook(copy_path, read_only=True)
copy_ws = copy_wb.active

print("Original rows:", src_ws.max_row)
print("Copy rows:", copy_ws.max_row)
print("Original cols:", src_ws.max_column)
print("Copy cols:", copy_ws.max_column)

for row_idx in range(1, 6):
    src_row = [src_ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 6)]
    copy_row = [copy_ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 6)]
    print(f"Row {row_idx} src : {src_row}")
    print(f"Row {row_idx} copy: {copy_row}")

status_src = [src_ws.cell(row=r, column=14).value for r in range(1, 12)]
status_copy = [copy_ws.cell(row=r, column=14).value for r in range(1, 12)]
print("Status src  :", status_src)
print("Status copy :", status_copy)
